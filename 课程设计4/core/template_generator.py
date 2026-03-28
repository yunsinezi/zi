# ============================================================
# core/template_generator.py — 型值表 Excel 模板生成器
#
# 阶段2 新增模块
# 生成标准格式的型值表 Excel 模板，供用户下载填写
# ============================================================

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os


def generate_template(output_path: str, example: bool = True) -> str:
    """
    生成标准型值表 Excel 模板文件。

    模板格式：
        第1行：标题（说明）
        第2行：列标题 → [站号, X坐标(m), WL0.0, WL0.5, WL1.0, ..., WL8.0]
        第3行起：数据行（如果 example=True，填入示例数据）
        最后几行：填写说明

    参数：
        output_path : str   输出文件路径
        example     : bool  是否填入示例数据（默认 True）

    返回：
        str  实际保存的文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "型值表"

    # ── 样式定义 ──────────────────────────────────────────
    # 标题行
    title_fill = PatternFill("solid", fgColor="1a4a8a")
    title_font = Font(name="宋体", size=13, bold=True, color="FFFFFF")

    # 列标题行
    header_fill = PatternFill("solid", fgColor="2e7dd1")
    header_font = Font(name="宋体", size=10, bold=True, color="FFFFFF")

    # 站号列（固定列）
    fixed_fill = PatternFill("solid", fgColor="e8f0fb")
    fixed_font = Font(name="宋体", size=10, bold=True, color="1a4a8a")

    # 数据单元格
    data_font   = Font(name="Courier New", size=10)
    data_fill   = PatternFill("solid", fgColor="FFFFFF")

    # 说明行
    note_fill = PatternFill("solid", fgColor="fffbf0")
    note_font = Font(name="宋体", size=9, color="856404", italic=True)

    # 边框
    thin  = Side(style="thin",   color="c8d8f0")
    thick = Side(style="medium", color="2e7dd1")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ── 水线高度列表（标准课设水线）────────────────────────
    # 从基线 0.0 到 8.0，每 0.5m 一条水线，共 17 条
    waterlines = [round(i * 0.5, 1) for i in range(17)]  # 0.0, 0.5, ..., 8.0

    # ── 示例站号（11站，站距1，中站x=0）────────────────────
    # 注意：这里站距=1，实际使用时乘以实际站距即可
    stations = list(range(-5, 6))  # -5, -4, ..., 0, ..., 5

    # ── 示例半宽数据（模拟典型货船型线）────────────────────
    import math
    def example_half_breadth(x, z, B_half=7.6, D=8.0):
        """生成示例半宽（模拟典型货船）"""
        x_norm = x / 5.0
        z_norm = z / D
        if abs(x_norm) > 1.0:
            return 0.0
        # 纵向：中部饱满，首尾收缩
        bf = 1.0 - 0.55 * (abs(x_norm) ** 2.8)
        # 垂向：随吃水增大
        df = 0.25 + 0.75 * (z_norm ** 0.35) if z_norm > 0 else 0.0
        y = B_half * bf * df
        # 端点强制为0
        if abs(x) == 5:
            y = 0.0
        return max(0.0, round(y, 3))

    # ── 第1行：大标题 ────────────────────────────────────
    total_cols = 2 + len(waterlines)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = "船舶静力学课程设计 — 型值表（半宽，单位：米）"
    ws["A1"].font      = title_font
    ws["A1"].fill      = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    # ── 第2行：列标题 ────────────────────────────────────
    # A列：站号
    c = ws.cell(2, 1, "站号")
    c.font = header_font; c.fill = header_fill
    c.alignment = center; c.border = header_border

    # B列：X坐标
    c = ws.cell(2, 2, "X坐标(m)")
    c.font = header_font; c.fill = header_fill
    c.alignment = center; c.border = header_border

    # C列起：各水线
    for j, z in enumerate(waterlines):
        col = j + 3
        label = f"WL{z:.1f}"
        c = ws.cell(2, col, label)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = header_border

    ws.row_dimensions[2].height = 22

    # ── 第3行起：数据行 ──────────────────────────────────
    for i, x in enumerate(stations):
        row = i + 3

        # 站号列
        c = ws.cell(row, 1, i + 1)  # 站号编号（1, 2, 3...）
        c.font = fixed_font; c.fill = fixed_fill
        c.alignment = center; c.border = data_border

        # X坐标列
        c = ws.cell(row, 2, x)
        c.font = fixed_font; c.fill = fixed_fill
        c.alignment = center; c.border = data_border
        c.number_format = "0.000"

        # 各水线半宽
        for j, z in enumerate(waterlines):
            col = j + 3
            if example:
                y = example_half_breadth(x, z)
            else:
                y = None  # 空白模板
            c = ws.cell(row, col, y)
            c.font = data_font; c.fill = data_fill
            c.alignment = center; c.border = data_border
            if y is not None:
                c.number_format = "0.000"

        ws.row_dimensions[row].height = 18

    # ── 说明区域 ─────────────────────────────────────────
    note_start = len(stations) + 3 + 2  # 数据行结束后空2行

    notes = [
        "【填写说明】",
        "1. 本模板为标准型值表格式，请按实际船型数据填写各水线处的半宽（单侧，单位：米）",
        "2. 坐标系：X轴沿船长方向，船首为正，船尾为负，中站 x=0",
        "3. 半宽 y 为单侧宽度，系统计算时自动乘以2得全宽",
        "4. 首尾端点（x最大和最小的站）半宽必须为0，系统会自动校验并修正",
        "5. 水线高度从基线（WL0.0）开始，向上递增，单位：米",
        "6. 如需增加水线，在右侧添加列，列标题格式为 WL{高度}，如 WL8.5",
        "7. 如需增加站，在数据区插入行，X坐标按实际填写（保持从小到大排列）",
        "8. 上传后系统会自动校验数据，给出校验报告",
    ]

    for k, note in enumerate(notes):
        row = note_start + k
        ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
        c = ws.cell(row, 1, note)
        c.font = note_font; c.fill = note_fill
        c.alignment = left
        ws.row_dimensions[row].height = 16

    # ── 列宽设置 ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 8   # 站号
    ws.column_dimensions["B"].width = 12  # X坐标
    for j in range(len(waterlines)):
        ws.column_dimensions[get_column_letter(j + 3)].width = 9

    # ── 冻结首行和前两列 ─────────────────────────────────
    ws.freeze_panes = "C3"

    # ── 保存 ─────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return output_path
