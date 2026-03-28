# ============================================================
# core/exporter_full.py — 多吃水静水力计算表 Excel 导出模块
#
# 阶段3 新增模块
# 导出完整的多吃水静水力计算表，格式符合课程设计要求
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def export_hydrostatics_excel(table: dict, output_path: str) -> str:
    """
    将多吃水静水力计算表导出为 Excel 文件。

    格式：
        第1行：大标题
        第2行：副标题（计算方法、日期）
        第3行：列标题（参数名）
        第4行：单位行
        第5行起：数据行（每行一个吃水）
        最后：公式说明区

    参数：
        table       : dict  calc_hydrostatics_table() 返回的结果
        output_path : str   输出文件路径

    返回：
        str  实际保存路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "静水力计算表"

    rows    = table["rows"]
    columns = table["columns"]
    method  = table.get("method", "trapz")
    method_name = "梯形法" if method == "trapz" else "辛普森1/3法"

    n_cols = len(columns)

    # ── 样式 ──────────────────────────────────────────────
    title_fill  = PatternFill("solid", fgColor="1a4a8a")
    title_font  = Font(name="宋体", size=13, bold=True, color="FFFFFF")
    sub_fill    = PatternFill("solid", fgColor="2e7dd1")
    sub_font    = Font(name="宋体", size=9,  bold=False, color="FFFFFF")
    hdr_fill    = PatternFill("solid", fgColor="e8f0fb")
    hdr_font    = Font(name="宋体", size=9,  bold=True,  color="1a4a8a")
    unit_fill   = PatternFill("solid", fgColor="f0f4fa")
    unit_font   = Font(name="宋体", size=8,  italic=True, color="555555")
    data_font   = Font(name="Courier New", size=9)
    alt_fill    = PatternFill("solid", fgColor="f8faff")
    note_fill   = PatternFill("solid", fgColor="fffbf0")
    note_font   = Font(name="宋体", size=8, italic=True, color="856404")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    thin   = Side(style="thin",   color="c8d8f0")
    medium = Side(style="medium", color="2e7dd1")
    tb     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hb     = Border(left=medium, right=medium, top=medium, bottom=medium)

    def sc(row, col, val, font=None, fill=None, align=None, border=None, nf=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:   c.font      = font
        if fill:   c.fill      = fill
        if align:  c.alignment = align
        if border: c.border    = border
        if nf:     c.number_format = nf
        return c

    # ── 第1行：大标题 ─────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    sc(1, 1, "船舶静水力计算表",
       font=title_font, fill=title_fill, align=center)
    ws.row_dimensions[1].height = 30

    # ── 第2行：副标题 ─────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sc(2, 1,
       f"武汉理工大学 船舶与海洋工程学院  |  计算方法：{method_name}  |  "
       f"依据：盛振邦《船舶原理》上册  |  "
       f"计算日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
       font=sub_font, fill=sub_fill, align=center)
    ws.row_dimensions[2].height = 18

    # ── 第3行：参数名 ─────────────────────────────────────
    for j, col in enumerate(columns):
        sc(3, j+1, col["label"],
           font=hdr_font, fill=hdr_fill, align=center, border=hb)
    ws.row_dimensions[3].height = 28

    # ── 第4行：单位 ───────────────────────────────────────
    for j, col in enumerate(columns):
        sc(4, j+1, col["unit"],
           font=unit_font, fill=unit_fill, align=center, border=tb)
    ws.row_dimensions[4].height = 16

    # ── 第5行起：数据 ─────────────────────────────────────
    for i, row in enumerate(rows):
        excel_row = i + 5
        row_fill = None if i % 2 == 0 else alt_fill

        for j, col in enumerate(columns):
            val = row.get(col["key"], "")
            c = sc(excel_row, j+1, val,
                   font=data_font,
                   fill=row_fill,
                   align=center,
                   border=tb,
                   nf="0.0000" if isinstance(val, float) else None)
        ws.row_dimensions[excel_row].height = 16

    # ── 公式说明区 ────────────────────────────────────────
    note_row = len(rows) + 6
    ws.merge_cells(f"A{note_row}:{get_column_letter(n_cols)}{note_row}")
    sc(note_row, 1, "【计算公式说明（对应盛振邦《船舶原理》上册）】",
       font=Font(name="宋体", size=9, bold=True, color="1a4a8a"),
       fill=PatternFill("solid", fgColor="e8f0fb"), align=left)
    ws.row_dimensions[note_row].height = 18

    formulas = [
        ("∇（排水体积）",    "∇ = ∫₀ᵈ Aw(z) dz",                    "各水线面积沿吃水方向积分"),
        ("Δ（排水量）",      "Δ = ρ·∇，海水ρ=1.025，淡水ρ=1.000",   "§2.1"),
        ("xb（浮心纵坐标）", "xb = (1/∇)·∫₀ᵈ Aw(z)·xf(z) dz",      "各水线面积形心的加权平均"),
        ("KB（浮心高度）",   "KB = (1/∇)·∫₀ᵈ Aw(z)·z dz",           "§2.2"),
        ("Aw（水线面积）",   "Aw = 2·∫y dx",                          "梯形法积分，y为半宽"),
        ("xf（漂心纵坐标）", "xf = (2·∫y·x dx) / Aw",                "水线面积形心"),
        ("TPC",              "TPC = Aw·ρ/100",                        "吃水增1cm的排水量增量（t/cm）"),
        ("MTC",              "MTC = ρ·∇·BML / (100·L)",              "使船纵倾1cm的力矩（t·m/cm）"),
        ("BM（横稳心半径）", "BM = IL/∇，IL=(2/3)·∫y³dx - Aw·xf²",  "§2.4 初稳性"),
        ("BML（纵稳心半径）","BML = IT/∇，IT=2·∫y·x²dx - Aw·xf²",   "纵稳性"),
        ("KM（稳心高度）",   "KM = KB + BM",                          "§2.4"),
        ("Cb（方形系数）",   "Cb = ∇/(L·B·d)",                       "§1.3 船型系数"),
        ("Cp（棱形系数）",   "Cp = ∇/(Am·L)",                        "Am为中横剖面面积"),
        ("Cwp（水线面系数）","Cwp = Aw/(L·B)",                        "§1.3"),
        ("Cm（中剖面系数）", "Cm = Am/(B·d)",                         "§1.3"),
    ]

    for k, (name, formula, note) in enumerate(formulas):
        r = note_row + 1 + k
        sc(r, 1, name,    font=note_font, fill=note_fill, align=left, border=tb)
        sc(r, 2, formula, font=Font(name="Courier New", size=8, color="1a4a8a"),
           fill=note_fill, align=left, border=tb)
        ws.merge_cells(f"B{r}:{get_column_letter(max(n_cols-1,2))}{r}")
        sc(r, n_cols, note, font=note_font, fill=note_fill, align=left, border=tb)
        ws.row_dimensions[r].height = 15

    # ── 列宽 ──────────────────────────────────────────────
    col_widths = {
        "d": 8, "V": 10, "delta_sea": 11, "delta_fresh": 11,
        "xb": 10, "zb": 10, "Aw": 10, "xf": 10,
        "TPC": 9, "MTC": 10, "BM": 9, "BML": 10, "zM": 9,
        "Cb": 8, "Cp": 8, "Cwp": 8, "Cm": 8,
    }
    for j, col in enumerate(columns):
        w = col_widths.get(col["key"], 10)
        ws.column_dimensions[get_column_letter(j+1)].width = w

    # ── 冻结首行和吃水列 ──────────────────────────────────
    ws.freeze_panes = "B5"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return output_path
