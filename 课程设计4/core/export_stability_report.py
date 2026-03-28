# ============================================================
# core/export_stability_report.py 鈥?绋虫€ф牎鏍告姤鍛?Excel 瀵煎嚭妯″潡
#
# 闃舵6 绗笁浼樺厛绾?#
# 鍔熻兘锛?#   1. 鐢熸垚瀹屾暣鐨勭ǔ鎬ф牎鏍告姤鍛?Excel
#   2. 鍖呭惈宸ュ喌鍙傛暟銆佽绠楃粨鏋溿€佽　鍑嗘寚鏍囥€佸垽瀹氱粨璁?#   3. 鏍煎紡瑙勮寖锛屽彲鐩存帴鐢ㄤ簬璇剧▼璁捐
#
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional
import os

# 鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲
# Excel 瀵煎嚭
# 鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲

class StabilityReportExporter:
    """绋虫€ф牎鏍告姤鍛?Excel 瀵煎嚭绫?""
    
    def __init__(self):
        """鍒濆鍖栧鍑哄櫒"""
        self.wb = None
        self.ws = None
    
    def create_single_condition_report(self, condition_name: str, 
                                      analysis_result: Dict,
                                      judgment_result: Dict,
                                      output_path: str) -> str:
        """
        鍒涘缓鍗曞伐鍐电ǔ鎬ф牎鏍告姤鍛?Excel
        
        鍙傛暟锛?            condition_name: 宸ュ喌鍚嶇О
            analysis_result: 鍒嗘瀽缁撴灉
            judgment_result: 鍒ゅ畾缁撴灉
            output_path: 杈撳嚭璺緞
        
        杩斿洖锛?            杈撳嚭鏂囦欢璺緞
        """
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "绋虫€ф牎鏍告姤鍛?
        
        # 璁剧疆鍒楀
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 20
        
        row = 1
        
        # 鈹€鈹€ 鏍囬 鈹€鈹€
        self.ws.merge_cells(f'A{row}:D{row}')
        title_cell = self.ws[f'A{row}']
        title_cell.value = '鑸硅埗绋虫€ф牎鏍告姤鍛?
        title_cell.font = Font(name='瀹嬩綋', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1a4a8a', end_color='1a4a8a', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 25
        row += 1
        
        # 鈹€鈹€ 宸ュ喌淇℃伅 鈹€鈹€
        row = self._add_section_header(row, '宸ュ喌淇℃伅')
        
        loading = analysis_result.get('loading', {})
        self._add_row(row, '宸ュ喌鍚嶇О', condition_name)
        row += 1
        self._add_row(row, '鎬婚噸閲?螖 (t)', f"{loading.get('total_weight', 0):.2f}")
        row += 1
        self._add_row(row, '閲嶅績绾靛悜 Xg (m)', f"{loading.get('xg', 0):.2f}")
        row += 1
        self._add_row(row, '閲嶅績妯悜 Yg (m)', f"{loading.get('yg', 0):.2f}")
        row += 1
        self._add_row(row, '閲嶅績鍨傚悜 Zg (m)', f"{loading.get('zg', 0):.2f}")
        row += 1
        
        # 鈹€鈹€ 娴€佽绠楃粨鏋?鈹€鈹€
        row = self._add_section_header(row, '娴€佽绠楃粨鏋?)
        
        floating_state = analysis_result.get('floating_state', {})
        self._add_row(row, '骞冲潎鍚冩按 (m)', f"{floating_state.get('draft_mean', 0):.2f}")
        row += 1
        self._add_row(row, '棣栧悆姘?(m)', f"{floating_state.get('draft_fwd', 0):.2f}")
        row += 1
        self._add_row(row, '灏惧悆姘?(m)', f"{floating_state.get('draft_aft', 0):.2f}")
        row += 1
        self._add_row(row, '绾靛€?(m)', f"{floating_state.get('trim', 0):.2f}")
        row += 1
        
        # 鈹€鈹€ 绋虫€ф寚鏍?鈹€鈹€
        row = self._add_section_header(row, '绋虫€ф寚鏍?)
        
        gz_curve = analysis_result.get('gz_curve', {})
        indicators = judgment_result.get('indicators', {})
        
        self._add_row(row, '鍒濈ǔ鎬ч珮搴?GM (m)', f"{gz_curve.get('GM', 0):.4f}")
        row += 1
        self._add_row(row, '鏈€澶у鍘熷姏鑷?GZ_max (m)', f"{indicators.get('GZ_max', 0):.4f}")
        row += 1
        self._add_row(row, '鏈€澶у鍘熷姏鑷傝搴?胃_max_gz (掳)', f"{indicators.get('theta_max_gz', 0):.1f}")
        row += 1
        self._add_row(row, '绋虫€ф秷澶辫 胃_vanish (掳)', f"{indicators.get('theta_vanish', 0):.1f}")
        row += 1
        self._add_row(row, '绋虫€ц　鍑嗘暟 K', f"{indicators.get('K', 0):.4f}")
        row += 1
        
        # 鈹€鈹€ 瑙勮寖琛″噯鏍℃牳 鈹€鈹€
        row = self._add_section_header(row, '瑙勮寖琛″噯鏍℃牳锛堛€婂浗鍐呰埅琛屾捣鑸规硶瀹氭楠屾妧鏈鍒欍€嬶級')
        
        judgments = judgment_result.get('judgments', {})
        
        # 琛ㄥご
        self.ws[f'A{row}'] = '琛″噯椤?
        self.ws[f'B{row}'] = '璁＄畻鍊?
        self.ws[f'C{row}'] = '瑙勮寖瑕佹眰'
        self.ws[f'D{row}'] = '鍒ゅ畾缁撴灉'
        
        for col in ['A', 'B', 'C', 'D']:
            cell = self.ws[f'{col}{row}']
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2e7dd1', end_color='2e7dd1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # 琛″噯椤规暟鎹?        criteria_names = {
            'GM': '鍒濈ǔ鎬ч珮搴?,
            'GZ_max': '鏈€澶у鍘熷姏鑷?,
            'theta_max_gz': '鏈€澶у鍘熷姏鑷傝搴?,
            'theta_vanish': '绋虫€ф秷澶辫',
            'K': '绋虫€ц　鍑嗘暟'
        }
        
        for key, name in criteria_names.items():
            if key in judgments:
                judgment = judgments[key]
                passed = judgment.get('passed', False)
                
                self.ws[f'A{row}'] = name
                self.ws[f'B{row}'] = f"{judgment.get('value', 0):.4f}"
                self.ws[f'C{row}'] = f"鈮?{judgment.get('limit', 0)}"
                self.ws[f'D{row}'] = '鉁?閫氳繃' if passed else '鉁?涓嶉€氳繃'
                
                # 涓嶉€氳繃鏃舵爣绾?                if not passed:
                    for col in ['A', 'B', 'C', 'D']:
                        self.ws[f'{col}{row}'].fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                
                row += 1
        
        # 鈹€鈹€ 鎬讳綋缁撹 鈹€鈹€
        row += 1
        row = self._add_section_header(row, '鎬讳綋缁撹')
        
        overall_pass = judgment_result.get('overall_pass', False)
        conclusion = '鉁?绋虫€у悎鏍? if overall_pass else '鉁?绋虫€т笉鍚堟牸'
        
        self.ws.merge_cells(f'A{row}:D{row}')
        conclusion_cell = self.ws[f'A{row}']
        conclusion_cell.value = conclusion
        conclusion_cell.font = Font(size=12, bold=True, color='FFFFFF')
        conclusion_cell.fill = PatternFill(
            start_color='27ae60' if overall_pass else 'e74c3c',
            end_color='27ae60' if overall_pass else 'e74c3c',
            fill_type='solid'
        )
        conclusion_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 20
        row += 1
        
        # 涓嶆弧瓒抽」
        if not overall_pass:
            failed_items = judgment_result.get('failed_items', [])
            if failed_items:
                self.ws.merge_cells(f'A{row}:D{row}')
                failed_cell = self.ws[f'A{row}']
                failed_cell.value = f'涓嶆弧瓒抽」锛歿", ".join(failed_items)}'
                failed_cell.font = Font(size=11, color='FFFFFF')
                failed_cell.fill = PatternFill(start_color='e74c3c', end_color='e74c3c', fill_type='solid')
                failed_cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
        
        # 鈹€鈹€ 椤佃剼 鈹€鈹€
        row += 2
        self.ws.merge_cells(f'A{row}:D{row}')
        footer_cell = self.ws[f'A{row}']
        footer_cell.value = f'鐢熸垚鏃堕棿锛歿datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        footer_cell.font = Font(size=9, italic=True, color='666666')
        footer_cell.alignment = Alignment(horizontal='right')
        
        # 淇濆瓨鏂囦欢
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        
        return output_path
    
    def create_all_conditions_report(self, all_results: Dict, 
                                    output_path: str) -> str:
        """
        鍒涘缓鍏ㄥ伐鍐电ǔ鎬ф牎鏍告姤鍛?Excel
        
        鍙傛暟锛?            all_results: {宸ュ喌鍚? {analysis, judgment}}
            output_path: 杈撳嚭璺緞
        
        杩斿洖锛?            杈撳嚭鏂囦欢璺緞
        """
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "鍏ㄥ伐鍐垫牎鏍?
        
        # 璁剧疆鍒楀
        for col in range(1, 10):
            self.ws.column_dimensions[get_column_letter(col)].width = 15
        
        row = 1
        
        # 鈹€鈹€ 鏍囬 鈹€鈹€
        self.ws.merge_cells(f'A{row}:I{row}')
        title_cell = self.ws[f'A{row}']
        title_cell.value = '鍏ㄥ伐鍐电ǔ鎬ф牎鏍告姤鍛?
        title_cell.font = Font(name='瀹嬩綋', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1a4a8a', end_color='1a4a8a', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 25
        row += 2
        
        # 鈹€鈹€ 琛ㄥご 鈹€鈹€
        headers = ['宸ュ喌鍚嶇О', '鍒ゅ畾缁撴灉', 'GM (m)', 'GZ_max (m)', '胃_max_gz (掳)', 
                  '胃_vanish (掳)', 'K 鍊?, '涓嶆弧瓒抽」', '瑙勮寖鏉℃']
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2e7dd1', end_color='2e7dd1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # 鈹€鈹€ 鏁版嵁琛?鈹€鈹€
        for condition_name, result in all_results.items():
            judgment = result.get('judgment', {})
            indicators = judgment.get('indicators', {})
            passed = judgment.get('overall_pass', False)
            failed_items = judgment.get('failed_items', [])
            
            self.ws.cell(row=row, column=1).value = condition_name
            self.ws.cell(row=row, column=2).value = '鉁?鍚堟牸' if passed else '鉁?涓嶅悎鏍?
            self.ws.cell(row=row, column=3).value = f"{indicators.get('GM', 0):.4f}"
            self.ws.cell(row=row, column=4).value = f"{indicators.get('GZ_max', 0):.4f}"
            self.ws.cell(row=row, column=5).value = f"{indicators.get('theta_max_gz', 0):.1f}"
            self.ws.cell(row=row, column=6).value = f"{indicators.get('theta_vanish', 0):.1f}"
            self.ws.cell(row=row, column=7).value = f"{indicators.get('K', 0):.4f}"
            self.ws.cell(row=row, column=8).value = ', '.join(failed_items) if failed_items else '鏃?
            self.ws.cell(row=row, column=9).value = '瑙勫垯绗?绡囩3绔?
            
            # 涓嶉€氳繃鏃舵爣绾?            if not passed:
                for col in range(1, 10):
                    self.ws.cell(row=row, column=col).fill = PatternFill(
                        start_color='ffcccc', end_color='ffcccc', fill_type='solid'
                    )
            
            row += 1
        
        # 鈹€鈹€ 鎬讳綋缁熻 鈹€鈹€
        row += 1
        total = len(all_results)
        passed_count = sum(1 for r in all_results.values() if r.get('judgment', {}).get('overall_pass', False))
        failed_count = total - passed_count
        
        self.ws.merge_cells(f'A{row}:I{row}')
        summary_cell = self.ws[f'A{row}']
        summary_cell.value = f'鎬昏锛歿total} 宸ュ喌 | 鍚堟牸锛歿passed_count} | 涓嶅悎鏍硷細{failed_count}'
        summary_cell.font = Font(size=11, bold=True, color='FFFFFF')
        summary_cell.fill = PatternFill(start_color='1a4a8a', end_color='1a4a8a', fill_type='solid')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 淇濆瓨鏂囦欢
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        
        return output_path
    
    def _add_section_header(self, row: int, title: str) -> int:
        """娣诲姞绔犺妭鏍囬"""
        self.ws.merge_cells(f'A{row}:D{row}')
        header_cell = self.ws[f'A{row}']
        header_cell.value = title
        header_cell.font = Font(bold=True, color='FFFFFF', size=11)
        header_cell.fill = PatternFill(start_color='2e7dd1', end_color='2e7dd1', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[row].height = 18
        return row + 1
    
    def _add_row(self, row: int, label: str, value: str):
        """娣诲姞鏁版嵁琛?""
        self.ws[f'A{row}'] = label
        self.ws[f'B{row}'] = value
        
        # 鏍峰紡
        for col in ['A', 'B']:
            cell = self.ws[f'{col}{row}']
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        self.ws[f'A{row}'].font = Font(bold=True)


# 鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲
# 娴嬭瘯鍑芥暟
# 鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲鈺愨晲

if __name__ == '__main__':
    import sys
    try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass
    
    print('=' * 70)
    print('绋虫€ф牎鏍告姤鍛?Excel 瀵煎嚭妯″潡娴嬭瘯')
    print('=' * 70)
    
    # 妯℃嫙鏁版嵁
    mock_analysis = {
        'condition_name': '婊¤浇鍑烘腐',
        'loading': {
            'total_weight': 5200,
            'xg': 42.5,
            'yg': 0,
            'zg': 5.1
        },
        'floating_state': {
            'draft_mean': 5.80,
            'draft_fwd': 5.80,
            'draft_aft': 5.80,
            'trim': 0.0
        },
        'gz_curve': {
            'GM': 0.8234
        }
    }
    
    mock_judgment = {
        'overall_pass': True,
        'indicators': {
            'GM': 0.8234,
            'GZ_max': 0.42,
            'theta_max_gz': 30.0,
            'theta_vanish': 85.0,
            'K': 1.2
        },
        'judgments': {
            'GM': {'passed': True, 'value': 0.8234, 'limit': 0.15},
            'GZ_max': {'passed': True, 'value': 0.42, 'limit': 0.20},
            'theta_max_gz': {'passed': True, 'value': 30.0, 'limit': 25.0},
            'theta_vanish': {'passed': True, 'value': 85.0, 'limit': 55.0},
            'K': {'passed': True, 'value': 1.2, 'limit': 1.0}
        },
        'failed_items': []
    }
    
    # 鍒涘缓瀵煎嚭鍣?    exporter = StabilityReportExporter()
    
    # 瀵煎嚭鍗曞伐鍐垫姤鍛?    print('\n銆愬鍑哄崟宸ュ喌鎶ュ憡銆?)
    output_path = 'F:/ship-statics/outputs/绋虫€ф牎鏍告姤鍛奯婊¤浇鍑烘腐.xlsx'
    exporter.create_single_condition_report('婊¤浇鍑烘腐', mock_analysis, mock_judgment, output_path)
    print(f'鉁?鍗曞伐鍐垫姤鍛婂凡瀵煎嚭锛歿output_path}')
    
    print('\n鉁?绋虫€ф牎鏍告姤鍛?Excel 瀵煎嚭妯″潡娴嬭瘯閫氳繃!')

