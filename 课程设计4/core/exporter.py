# ============================================================
# core/exporter.py — Excel 导出模块
#
# 使用 openpyxl 生成符合课程设计格式要求的 Excel 计算表
# 格式规范：标题、参数名称、数值、单位，工整清晰
# ============================================================

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def export_to_excel(result: dict, stations: list, half_breadths: list,
                    output_path: str) -> str:
    """
    将静水力计算结果导出为 Excel 文件
    
    格式要求（符合武理船海课程设计规范）：
        - 第一部分：封面信息（课程名称、计算日期）
        - 第二部分：基本参数（LBP、设计吃水等）
        - 第三部分：型值表（站号 + 半宽）
        - 第四部分：静水力计算结果（参数名、数值、单位、公式说明）
    
    参数：
        result      : dict  calculate_hydrostatics() 返回的计算结果
        stations    : list  各站 x 坐标
        half_breadths: list 各站半宽
        output_path : str   输出文件路径（.xlsx）
    
    返回：
        str  实际保存的文件路径
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "静水力计算"
    
    # ── 样式定义 ──────────────────────────────────────────
    # 标题样式（深蓝底白字）
    title_fill = PatternFill("solid", fgColor="1a4a8a")
    title_font = Font(name="宋体", size=14, bold=True, color="FFFFFF")
    
    # 节标题样式（浅蓝底深蓝字）
    section_fill = PatternFill("solid", fgColor="d0e4f8")
    section_font = Font(name="宋体", size=11, bold=True, color="1a4a8a")
    
    # 表头样式（中蓝底白字）
    header_fill = PatternFill("solid", fgColor="2e7dd1")
    header_font = Font(name="宋体", size=10, bold=True, color="FFFFFF")
    
    # 数据样式（普通）
    data_font = Font(name="宋体", size=10)
    
    # 强调样式（橙色底）
    accent_fill = PatternFill("solid", fgColor="fff3cd")
    accent_font = Font(name="宋体", size=10, bold=True, color="856404")
    
    # 居中对齐
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    
    # 细边框
    thin = Side(style="thin", color="aaaaaa")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # 粗边框（用于分节）
    thick = Side(style="medium", color="1a4a8a")
    thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    
    # ── 辅助函数：设置单元格样式 ──────────────────────────
    def set_cell(row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
        """设置单元格内容和样式"""
        cell = ws.cell(row=row, column=col, value=value)
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border
        if num_fmt: cell.number_format = num_fmt
        return cell
    
    # ── 第一部分：封面标题 ────────────────────────────────
    # 合并A1:F1作为大标题
    ws.merge_cells("A1:F1")
    set_cell(1, 1,
             "船舶静力学课程设计计算书",
             font=title_font, fill=title_fill, align=center_align)
    ws.row_dimensions[1].height = 36
    
    # 副标题
    ws.merge_cells("A2:F2")
    set_cell(2, 1,
             "武汉理工大学 船舶与海洋工程学院  |  计算方法：梯形法  |  依据：盛振邦《船舶原理》上册",
             font=Font(name="宋体", size=9, color="555555"),
             fill=PatternFill("solid", fgColor="f0f4fa"),
             align=center_align)
    
    # 计算日期
    ws.merge_cells("A3:F3")
    set_cell(3, 1,
             f"计算日期：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}",
             font=Font(name="宋体", size=9, color="888888"),
             align=Alignment(horizontal="right", vertical="center"))
    
    current_row = 5  # 从第5行开始写内容
    
    # ── 第二部分：基本参数 ────────────────────────────────
    ws.merge_cells(f"A{current_row}:F{current_row}")
    set_cell(current_row, 1, "一、基本参数",
             font=section_font, fill=section_fill, align=left_align)
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    
    # 表头
    headers = ["参数名称", "符号", "数值", "单位", "说明"]
    col_widths = [20, 10, 14, 10, 30]
    for j, (h, w) in enumerate(zip(headers, col_widths)):
        set_cell(current_row, j+1, h,
                 font=header_font, fill=header_fill,
                 align=center_align, border=thin_border)
        ws.column_dimensions[get_column_letter(j+1)].width = w
    current_row += 1
    
    # 基本参数数据行
    basic_params = [
        ("垂线间长",   "LBP",  result["L"],     "m",    "首垂线到尾垂线的距离"),
        ("型宽",       "B",    result["B"],     "m",    "设计吃水处最大型宽"),
        ("设计吃水",   "d",    result["d"],     "m",    "设计满载吃水"),
        ("海水密度",   "ρ",    1.025,           "t/m³", "按规范取值"),
    ]
    for name, sym, val, unit, note in basic_params:
        set_cell(current_row, 1, name, font=data_font, align=left_align, border=thin_border)
        set_cell(current_row, 2, sym,  font=data_font, align=center_align, border=thin_border)
        set_cell(current_row, 3, val,  font=data_font, align=center_align, border=thin_border, num_fmt="0.000")
        set_cell(current_row, 4, unit, font=data_font, align=center_align, border=thin_border)
        set_cell(current_row, 5, note, font=data_font, align=left_align,   border=thin_border)
        current_row += 1
    
    current_row += 1  # 空一行
    
    # ── 第三部分：型值表 ──────────────────────────────────
    ws.merge_cells(f"A{current_row}:F{current_row}")
    set_cell(current_row, 1, "二、设计吃水处型值表（半宽）",
             font=section_font, fill=section_fill, align=left_align)
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    
    # 坐标系说明
    ws.merge_cells(f"A{current_row}:F{current_row}")
    set_cell(current_row, 1,
             "坐标系：X轴沿船长方向（船首为正，船尾为负，中站x=0），y为半宽（单侧）",
             font=Font(name="宋体", size=9, color="666666", italic=True),
             align=left_align)
    current_row += 1
    
    # 型值表表头
    set_cell(current_row, 1, "站号 x (m)",
             font=header_font, fill=header_fill, align=center_align, border=thin_border)
    set_cell(current_row, 2, "半宽 y (m)",
             font=header_font, fill=header_fill, align=center_align, border=thin_border)
    set_cell(current_row, 3, "2y (全宽, m)",
             font=header_font, fill=header_fill, align=center_align, border=thin_border)
    set_cell(current_row, 4, "y·x (m²)",
             font=header_font, fill=header_fill, align=center_align, border=thin_border)
    current_row += 1
    
    # 型值表数据
    for x, y in zip(stations, half_breadths):
        set_cell(current_row, 1, x,       font=data_font, align=center_align, border=thin_border, num_fmt="0.000")
        set_cell(current_row, 2, y,       font=data_font, align=center_align, border=thin_border, num_fmt="0.000")
        set_cell(current_row, 3, 2*y,     font=data_font, align=center_align, border=thin_border, num_fmt="0.000")
        set_cell(current_row, 4, y*x,     font=data_font, align=center_align, border=thin_border, num_fmt="0.000")
        current_row += 1
    
    current_row += 1  # 空一行
    
    # ── 第四部分：静水力计算结果 ──────────────────────────
    ws.merge_cells(f"A{current_row}:F{current_row}")
    set_cell(current_row, 1, "三、静水力要素计算结果（设计吃水）",
             font=section_font, fill=section_fill, align=left_align)
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    
    # 结果表头
    res_headers = ["计算量", "符号", "数值", "单位", "计算公式"]
    for j, h in enumerate(res_headers):
        set_cell(current_row, j+1, h,
                 font=header_font, fill=header_fill,
                 align=center_align, border=thin_border)
    current_row += 1
    
    # 计算结果数据
    results_data = [
        ("水线面积",       "Aw",  result["Aw"],    "m²",   "Aw = 2·∫y dx  （梯形法）"),
        ("排水体积",       "∇",   result["V"],     "m³",   "∇ = Aw · d  （简化）"),
        ("排水量",         "Δ",   result["delta"], "t",    "Δ = ρ·∇，ρ=1.025 t/m³"),
        ("浮心纵向坐标",   "xb",  result["xb"],    "m",    "xb = (1/Aw)·2·∫y·x dx"),
        ("每厘米吃水吨数", "TPC", result["TPC"],   "t/cm", "TPC = Aw·ρ/100"),
        ("方形系数",       "Cb",  result["Cb"],    "—",    "Cb = ∇/(L·B·d)"),
    ]
    for name, sym, val, unit, formula in results_data:
        set_cell(current_row, 1, name,    font=accent_font, fill=accent_fill, align=left_align,   border=thin_border)
        set_cell(current_row, 2, sym,     font=accent_font, fill=accent_fill, align=center_align, border=thin_border)
        set_cell(current_row, 3, val,     font=Font(name="宋体", size=10, bold=True, color="c0392b"),
                 fill=accent_fill, align=center_align, border=thin_border, num_fmt="0.0000")
        set_cell(current_row, 4, unit,    font=accent_font, fill=accent_fill, align=center_align, border=thin_border)
        set_cell(current_row, 5, formula, font=data_font,   align=left_align,   border=thin_border)
        current_row += 1
    
    # ── 设置行高 ──────────────────────────────────────────
    for row in ws.iter_rows():
        if ws.row_dimensions[row[0].row].height is None:
            ws.row_dimensions[row[0].row].height = 18
    
    # ── 保存文件 ──────────────────────────────────────────
    wb.save(output_path)
    return output_path
