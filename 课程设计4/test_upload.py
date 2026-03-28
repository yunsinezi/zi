"""
文件上传测试脚本
用于测试 Excel 文件上传功能是否正常工作
"""

import requests
import os
import sys

# 测试配置
BASE_URL = "http://127.0.0.1:5000"
TEST_FILE = "test_template.xlsx"

def test_backend_running():
    """测试后端是否运行"""
    print("\n【测试 1】检查后端是否运行...")
    try:
        response = requests.get(f"{BASE_URL}/", timeout=5)
        if response.status_code == 200:
            print("✓ 后端正在运行")
            return True
        else:
            print(f"✗ 后端返回状态码：{response.status_code}")
            return False
    except requests.exceptions.ConnectionError:
        print("✗ 无法连接到后端，请确认 Flask 已启动")
        return False
    except Exception as e:
        print(f"✗ 测试失败：{e}")
        return False

def test_template_download():
    """测试模板下载"""
    print("\n【测试 2】测试模板下载...")
    try:
        response = requests.get(f"{BASE_URL}/download_template", timeout=10)
        if response.status_code == 200:
            # 保存模板
            with open("downloaded_template.xlsx", "wb") as f:
                f.write(response.content)
            print("✓ 模板下载成功")
            print(f"  文件大小：{len(response.content) / 1024:.2f} KB")
            return True
        else:
            print(f"✗ 下载失败，状态码：{response.status_code}")
            return False
    except Exception as e:
        print(f"✗ 测试失败：{e}")
        return False

def test_file_upload(file_path):
    """测试文件上传"""
    print(f"\n【测试 3】测试文件上传：{file_path}...")
    
    if not os.path.exists(file_path):
        print(f"✗ 文件不存在：{file_path}")
        return False
    
    try:
        # 准备上传
        file_size = os.path.getsize(file_path) / 1024
        print(f"  文件大小：{file_size:.2f} KB")
        
        # 发送上传请求
        with open(file_path, "rb") as f:
            files = {"file": (os.path.basename(file_path), f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            response = requests.post(f"{BASE_URL}/upload_offsets", files=files, timeout=30)
        
        # 检查响应
        if response.status_code == 200:
            data = response.json()
            if data.get("success"):
                print("✓ 上传成功")
                
                # 显示解析结果
                result_data = data.get("data", {})
                print(f"  站数：{result_data.get('n_stations', 'N/A')}")
                print(f"  水线数：{result_data.get('n_waterlines', 'N/A')}")
                
                # 显示校验结果
                validation = result_data.get("validation", {})
                if validation.get("is_valid"):
                    print("  校验：✓ 通过")
                else:
                    print("  校验：✗ 未通过")
                    if validation.get("errors"):
                        for err in validation["errors"]:
                            print(f"    错误：{err}")
                
                return True
            else:
                print(f"✗ 上传失败：{data.get('msg', '未知错误')}")
                return False
        else:
            print(f"✗ 上传失败，状态码：{response.status_code}")
            print(f"  响应内容：{response.text[:200]}")
            return False
            
    except requests.exceptions.ConnectionError:
        print("✗ 无法连接到服务器")
        return False
    except requests.exceptions.Timeout:
        print("✗ 请求超时")
        return False
    except Exception as e:
        print(f"✗ 测试失败：{e}")
        import traceback
        traceback.print_exc()
        return False

def test_cors():
    """测试 CORS 支持"""
    print("\n【测试 4】测试 CORS 支持...")
    try:
        # 发送 OPTIONS 请求
        response = requests.options(
            f"{BASE_URL}/upload_offsets",
            headers={
                "Origin": "http://localhost:3000",
                "Access-Control-Request-Method": "POST"
            },
            timeout=5
        )
        
        # 检查 CORS 头
        cors_header = response.headers.get("Access-Control-Allow-Origin")
        if cors_header:
            print(f"✓ CORS 已启用，允许来源：{cors_header}")
            return True
        else:
            print("⚠ CORS 头未设置，可能会导致跨域问题")
            return False
    except Exception as e:
        print(f"⚠ CORS 测试失败：{e}")
        return False

def create_test_file():
    """创建测试文件"""
    print("\n【创建测试文件】")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "型值表"
        
        # 表头
        headers = ["站号", "水线0", "水线1", "水线2", "水线3", "水线4", "水线5"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 数据
        data = [
            [-5, 0.00, 0.50, 1.00, 1.50, 2.00, 2.50],
            [-4, 1.85, 2.50, 3.20, 4.00, 4.80, 5.50],
            [-3, 4.62, 5.50, 6.30, 7.00, 7.60, 8.10],
            [-2, 6.48, 7.30, 8.00, 8.60, 9.10, 9.50],
            [-1, 7.35, 8.10, 8.70, 9.20, 9.60, 9.90],
            [0,  7.60, 8.30, 8.90, 9.40, 9.80, 10.10],
            [1,  7.35, 8.10, 8.70, 9.20, 9.60, 9.90],
            [2,  6.48, 7.30, 8.00, 8.60, 9.10, 9.50],
            [3,  4.62, 5.50, 6.30, 7.00, 7.60, 8.10],
            [4,  1.85, 2.50, 3.20, 4.00, 4.80, 5.50],
            [5,  0.00, 0.50, 1.00, 1.50, 2.00, 2.50],
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)
        
        # 保存
        wb.save(TEST_FILE)
        print(f"✓ 测试文件已创建：{TEST_FILE}")
        return True
        
    except ImportError:
        print("✗ 未安装 openpyxl，无法创建测试文件")
        print("  请运行：pip install openpyxl")
        return False
    except Exception as e:
        print(f"✗ 创建测试文件失败：{e}")
        return False

def main():
    print("=" * 70)
    print("文件上传功能测试")
    print("=" * 70)
    
    # 检查依赖
    try:
        import flask_cors
        print("✓ flask-cors 已安装")
    except ImportError:
        print("✗ flask-cors 未安装，请运行：pip install flask-cors")
        return
    
    # 运行测试
    results = []
    
    # 测试 1：后端运行
    results.append(("后端运行", test_backend_running()))
    
    # 测试 2：模板下载
    results.append(("模板下载", test_template_download()))
    
    # 创建测试文件
    if not os.path.exists(TEST_FILE):
        create_test_file()
    
    # 测试 3：文件上传
    if os.path.exists(TEST_FILE):
        results.append(("文件上传", test_file_upload(TEST_FILE)))
    else:
        results.append(("文件上传", False))
    
    # 测试 4：CORS
    results.append(("CORS 支持", test_cors()))
    
    # 总结
    print("\n" + "=" * 70)
    print("测试结果汇总")
    print("=" * 70)
    
    for name, result in results:
        status = "✓ 通过" if result else "✗ 失败"
        print(f"{name:20s} {status}")
    
    # 统计
    passed = sum(1 for _, r in results if r)
    total = len(results)
    print(f"\n通过率：{passed}/{total} ({passed/total*100:.0f}%)")
    
    # 给出建议
    if passed == total:
        print("\n✓ 所有测试通过！文件上传功能正常。")
    else:
        print("\n⚠ 部分测试失败，请检查：")
        if not results[0][1]:  # 后端运行
            print("  1. Flask 后端未启动，请运行：python app.py")
        if not results[4][1]:  # CORS
            print("  2. CORS 未启用，请确保 app.py 中已添加 CORS(app)")
        if not results[2][1]:  # 文件上传
            print("  3. 文件上传失败，请检查接口地址和文件格式")

if __name__ == "__main__":
    main()
